import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.utils import timezone
from django.db.models import Q
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from .models import PersonalPolicial, PermisoLicencia, SancionAplicada, FelicitacionAplicada, KardexDigital, DestinoPolicial

from .models import (
    PersonalPolicial,
    PermisoLicencia,
    SancionAplicada,
    FelicitacionAplicada,
    KardexDigital
)
from catalogos.models import TipoFelicitacion

from core.mixins import (
    AdminRequiredMixin,
    OficialAdministrativoRequiredMixin,
    UsuarioAutorizadoRequiredMixin,
    PuedeCrearMixin,
    PuedeEditarMixin,
    PuedeEliminarMixin,
    PuedeAprobarPermisosMixin,
    PuedeGestionarSancionesMixin
)

from reportes.utils import BitacoraMixin, registrar_log


# ==========================================
# VISTAS PARA PERSONAL
# ==========================================

class PersonalListView(UsuarioAutorizadoRequiredMixin, ListView):
    model = PersonalPolicial
    template_name = 'personal/personal_list.html'
    context_object_name = 'personal'
    paginate_by = 20

    def get_queryset(self):
        queryset = PersonalPolicial.objects.select_related(
            'grado', 'unidad', 'estado_actual'
        )

        buscar = self.request.GET.get('buscar', '').strip()
        if buscar:
            queryset = queryset.filter(
                Q(nombres__icontains=buscar)          |
                Q(apellido_paterno__icontains=buscar) |
                Q(apellido_materno__icontains=buscar) |
                Q(ci__icontains=buscar)               |
                Q(codigo_identificacion__icontains=buscar)
            )

        grado_id = self.request.GET.get('grado', '')
        if grado_id:
            queryset = queryset.filter(grado_id=grado_id)

        unidad_id = self.request.GET.get('unidad', '')
        if unidad_id:
            queryset = queryset.filter(unidad_id=unidad_id)

        estado_id = self.request.GET.get('estado', '')
        if estado_id:
            queryset = queryset.filter(estado_actual_id=estado_id)

        genero = self.request.GET.get('genero', '')
        if genero:
            queryset = queryset.filter(genero=genero)

        activo = self.request.GET.get('activo', '')
        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset.order_by('grado__orden', 'apellido_paterno', 'nombres')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['buscar']     = self.request.GET.get('buscar', '')
        context['grado_sel']  = self.request.GET.get('grado', '')
        context['unidad_sel'] = self.request.GET.get('unidad', '')
        context['estado_sel'] = self.request.GET.get('estado', '')
        context['genero_sel'] = self.request.GET.get('genero', '')
        context['activo_sel'] = self.request.GET.get('activo', '')

        from catalogos.models import Grado, Unidad, TipoEstado
        context['grados']         = Grado.objects.filter(activo=True).order_by('orden')
        context['unidades']       = Unidad.objects.filter(activa=True).order_by('nombre')
        context['estados']        = TipoEstado.objects.all()
        context['total_general']  = PersonalPolicial.objects.count()
        return context


class PersonalCreateView(BitacoraMixin, PuedeCrearMixin, CreateView):
    bitacora_modulo = 'personal'
    model = PersonalPolicial
    template_name = 'personal/personal_form.html'
    fields = [
        'codigo_identificacion', 'ci', 'expedido',
        'nombres', 'apellido_paterno', 'apellido_materno',
        'fecha_nacimiento', 'genero',
        'grado', 'unidad', 'estado_actual', 'fecha_ingreso',
        'cargo_actual',
        'telefono_personal', 'telefono_emergencia',
        'correo_institucional', 'direccion_domicilio',
        'otra_profesion',
        'activo',
    ]
    success_url = reverse_lazy('personal_list')


class PersonalUpdateView(BitacoraMixin, PuedeEditarMixin, UpdateView):
    bitacora_modulo = 'personal'
    model = PersonalPolicial
    template_name = 'personal/personal_form.html'
    fields = [
        'codigo_identificacion', 'ci', 'expedido',
        'nombres', 'apellido_paterno', 'apellido_materno',
        'fecha_nacimiento', 'genero',
        'grado', 'unidad', 'estado_actual', 'fecha_ingreso',
        'cargo_actual',
        'telefono_personal', 'telefono_emergencia',
        'correo_institucional', 'direccion_domicilio',
        'otra_profesion',
        'activo',
    ]
    success_url = reverse_lazy('personal_list')


class PersonalDetailView(UsuarioAutorizadoRequiredMixin, DetailView):
    model = PersonalPolicial
    template_name = 'personal/personal_detail.html'
    context_object_name = 'persona'


class PersonalDeleteView(BitacoraMixin, PuedeEliminarMixin, DeleteView):
    bitacora_modulo = 'personal'
    model = PersonalPolicial
    template_name = 'personal/personal_confirm_delete.html'
    success_url = reverse_lazy('personal_list')


# ==========================================
# VISTAS PARA KARDEX
# ==========================================

class KardexPersonalListView(UsuarioAutorizadoRequiredMixin, ListView):
    model = KardexDigital
    template_name = 'personal/kardex_list.html'
    context_object_name = 'registros'

    def get_queryset(self):
        personal_id = self.kwargs['personal_id']
        return KardexDigital.objects.filter(personal_id=personal_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['persona'] = PersonalPolicial.objects.get(pk=self.kwargs['personal_id'])
        return context


class KardexCreateView(BitacoraMixin, PuedeCrearMixin, CreateView):
    bitacora_modulo = 'kardex'
    model = KardexDigital
    template_name = 'personal/kardex_form.html'
    fields = [
        'tipo_registro',
        'fecha_registro',
        'descripcion',
        'documento_referencia',
        'observaciones',
        'grado_anterior', 'grado_nuevo',
        'unidad_anterior', 'unidad_nueva',
        'estado_anterior', 'estado_nuevo',
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['grado_anterior'].required  = False
        form.fields['grado_nuevo'].required     = False
        form.fields['unidad_anterior'].required = False
        form.fields['unidad_nueva'].required    = False
        form.fields['estado_anterior'].required = False
        form.fields['estado_nuevo'].required    = False

        personal = PersonalPolicial.objects.get(pk=self.kwargs['personal_id'])
        form.fields['grado_anterior'].initial  = personal.grado
        form.fields['unidad_anterior'].initial = personal.unidad
        form.fields['estado_anterior'].initial = personal.estado_actual
        return form

    def form_valid(self, form):
        personal_id = self.kwargs['personal_id']
        form.instance.personal_id    = personal_id
        form.instance.registrado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('kardex_list', kwargs={'personal_id': self.kwargs['personal_id']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['persona'] = PersonalPolicial.objects.get(pk=self.kwargs['personal_id'])
        return context


# ==========================================
# VISTAS PARA PERMISOS
# ==========================================

class PermisoListView(UsuarioAutorizadoRequiredMixin, ListView):
    model = PermisoLicencia
    template_name = 'personal/permiso_list.html'
    context_object_name = 'permisos'
    ordering = ['-fecha_solicitud']


class PermisoCreateView(BitacoraMixin, PuedeCrearMixin, CreateView):
    bitacora_modulo = 'permisos'
    model = PermisoLicencia
    template_name = 'personal/permiso_form.html'
    fields = [
        'personal', 'tipo_permiso', 'fecha_inicio', 'fecha_fin',
        'motivo', 'documento_adjunto', 'numero_oficio'
    ]
    success_url = reverse_lazy('permiso_list')

    def form_valid(self, form):
        form.instance.solicitado_por = self.request.user
        return super().form_valid(form)


class PermisoUpdateView(BitacoraMixin, PuedeEditarMixin, UpdateView):
    bitacora_modulo = 'permisos'
    model = PermisoLicencia
    template_name = 'personal/permiso_form.html'
    fields = [
        'personal', 'tipo_permiso', 'fecha_inicio', 'fecha_fin',
        'motivo', 'documento_adjunto', 'numero_oficio'
    ]
    success_url = reverse_lazy('permiso_list')


class PermisoAprobarView(PuedeAprobarPermisosMixin, UpdateView):
    model = PermisoLicencia
    template_name = 'personal/permiso_aprobar.html'
    fields = ['estado', 'observaciones_aprobacion']

    def form_valid(self, form):
        if form.cleaned_data['estado'] == 'aprobado':
            form.instance.aprobado_por      = self.request.user
            form.instance.fecha_aprobacion  = timezone.now()

        # Log manual porque la acción es "aprobar", no un CRUD estándar
        registrar_log(
            self.request, 'EDITAR', 'permisos',
            f'Cambió estado de permiso a "{form.cleaned_data["estado"]}" '
            f'— {form.instance.personal}',
            objeto=form.instance,
        )
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('permiso_list')


class PermisoDetailView(UsuarioAutorizadoRequiredMixin, DetailView):
    model = PermisoLicencia
    template_name = 'personal/permiso_detail.html'
    context_object_name = 'permiso'


# ==========================================
# VISTAS PARA SANCIONES
# ==========================================

class SancionListView(PuedeGestionarSancionesMixin, ListView):
    model = SancionAplicada
    template_name = 'personal/sancion_list.html'
    context_object_name = 'sanciones'
    ordering = ['-fecha_sancion']


class SancionCreateView(BitacoraMixin, PuedeGestionarSancionesMixin, CreateView):
    bitacora_modulo = 'sanciones'
    model = SancionAplicada
    template_name = 'personal/sancion_form.html'
    fields = [
        'personal', 'tipo_sancion', 'fecha_sancion', 'fecha_inicio', 'fecha_fin',
        'motivo', 'estado', 'observaciones', 'documento_referencia'
    ]
    success_url = reverse_lazy('sancion_list')

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user
        return super().form_valid(form)


class SancionUpdateView(BitacoraMixin, PuedeGestionarSancionesMixin, UpdateView):
    bitacora_modulo = 'sanciones'
    model = SancionAplicada
    template_name = 'personal/sancion_form.html'
    fields = [
        'personal', 'tipo_sancion', 'fecha_sancion', 'fecha_inicio', 'fecha_fin',
        'motivo', 'estado', 'observaciones', 'documento_referencia'
    ]
    success_url = reverse_lazy('sancion_list')


class SancionDetailView(PuedeGestionarSancionesMixin, DetailView):
    model = SancionAplicada
    template_name = 'personal/sancion_detail.html'
    context_object_name = 'sancion'


# ==========================================
# VISTAS PARA FELICITACIONES
# ==========================================

class FelicitacionListView(UsuarioAutorizadoRequiredMixin, ListView):
    model = FelicitacionAplicada
    template_name = 'personal/felicitacion_list.html'
    context_object_name = 'felicitaciones'
    ordering = ['-fecha_felicitacion']

    def get_queryset(self):
        queryset = super().get_queryset()
        tipo_id = self.request.GET.get('tipo')
        if tipo_id:
            queryset = queryset.filter(tipo_felicitacion_id=tipo_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipos_felicitacion'] = TipoFelicitacion.objects.all()
        return context


class FelicitacionCreateView(BitacoraMixin, PuedeCrearMixin, CreateView):
    bitacora_modulo = 'felicitaciones'
    model = FelicitacionAplicada
    template_name = 'personal/felicitacion_form.html'
    fields = [
        'personal', 'tipo_felicitacion', 'fecha_felicitacion',
        'motivo', 'documento_referencia', 'observaciones'
    ]
    success_url = reverse_lazy('felicitacion_list')

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user
        return super().form_valid(form)


class FelicitacionUpdateView(BitacoraMixin, PuedeEditarMixin, UpdateView):
    bitacora_modulo = 'felicitaciones'
    model = FelicitacionAplicada
    template_name = 'personal/felicitacion_form.html'
    fields = [
        'personal', 'tipo_felicitacion', 'fecha_felicitacion',
        'motivo', 'documento_referencia', 'observaciones'
    ]
    success_url = reverse_lazy('felicitacion_list')


class FelicitacionDetailView(UsuarioAutorizadoRequiredMixin, DetailView):
    model = FelicitacionAplicada
    template_name = 'personal/felicitacion_detail.html'
    context_object_name = 'felicitacion'


# ==========================================
# VISTAS PARA DESTINOS
# ==========================================

class DestinoListView(UsuarioAutorizadoRequiredMixin, ListView):
    model = DestinoPolicial
    template_name = 'personal/destino_list.html'
    context_object_name = 'destinos'
    paginate_by = 20

    def get_queryset(self):
        queryset = DestinoPolicial.objects.select_related(
            'personal', 'personal__grado', 'unidad_destino'
        )
        personal_id = self.request.GET.get('personal')
        if personal_id:
            queryset = queryset.filter(personal_id=personal_id)

        estado = self.request.GET.get('estado', '')
        if estado == 'activo':
            queryset = queryset.filter(activo=True)
        elif estado == 'inactivo':
            queryset = queryset.filter(activo=False)

        buscar = self.request.GET.get('buscar', '')
        if buscar:
            queryset = queryset.filter(
                Q(personal__nombres__icontains=buscar)         |
                Q(personal__apellido_paterno__icontains=buscar)|
                Q(lugar_destino__icontains=buscar)             |
                Q(numero_resolucion__icontains=buscar)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['buscar']        = self.request.GET.get('buscar', '')
        context['estado']        = self.request.GET.get('estado', '')
        context['total_activos'] = DestinoPolicial.objects.filter(activo=True).count()
        return context


class DestinoPersonalListView(UsuarioAutorizadoRequiredMixin, ListView):
    model = DestinoPolicial
    template_name = 'personal/destino_personal_list.html'
    context_object_name = 'destinos'

    def get_queryset(self):
        return DestinoPolicial.objects.filter(
            personal_id=self.kwargs['personal_id']
        ).select_related('unidad_destino')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['persona'] = get_object_or_404(PersonalPolicial, pk=self.kwargs['personal_id'])
        return context


class DestinoCreateView(BitacoraMixin, PuedeCrearMixin, CreateView):
    bitacora_modulo = 'destinos'
    model = DestinoPolicial
    template_name = 'personal/destino_form.html'
    fields = [
        'personal', 'tipo_destino', 'unidad_destino', 'lugar_destino',
        'fecha_inicio', 'fecha_fin', 'activo',
        'descripcion', 'numero_resolucion', 'observaciones',
    ]

    def get_initial(self):
        initial = super().get_initial()
        personal_id = self.kwargs.get('personal_id') or self.request.GET.get('personal')
        if personal_id:
            initial['personal'] = personal_id
        return initial

    def form_valid(self, form):
        form.instance.registrado_por = self.request.user
        if form.instance.activo:
            DestinoPolicial.objects.filter(
                personal=form.instance.personal,
                activo=True
            ).update(activo=False)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('destino_personal_list',
                            kwargs={'personal_id': self.object.personal_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        personal_id = self.kwargs.get('personal_id') or self.request.GET.get('personal')
        if personal_id:
            context['persona'] = get_object_or_404(PersonalPolicial, pk=personal_id)
        return context


class DestinoUpdateView(BitacoraMixin, PuedeEditarMixin, UpdateView):
    bitacora_modulo = 'destinos'
    model = DestinoPolicial
    template_name = 'personal/destino_form.html'
    fields = [
        'personal', 'tipo_destino', 'unidad_destino', 'lugar_destino',
        'fecha_inicio', 'fecha_fin', 'activo',
        'descripcion', 'numero_resolucion', 'observaciones',
    ]

    def get_success_url(self):
        return reverse_lazy('destino_personal_list',
                            kwargs={'personal_id': self.object.personal_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['persona'] = self.object.personal
        return context


class DestinoDetailView(UsuarioAutorizadoRequiredMixin, DetailView):
    model = DestinoPolicial
    template_name = 'personal/destino_detail.html'
    context_object_name = 'destino'


class DestinoDeleteView(BitacoraMixin, PuedeEliminarMixin, DeleteView):
    bitacora_modulo = 'destinos'
    model = DestinoPolicial
    template_name = 'personal/destino_confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('destino_personal_list',
                            kwargs={'personal_id': self.object.personal_id})


# ==========================================
# REPORTE EXCEL
# ==========================================

def _build_workbook(queryset):
    HEADERS = [
        "GRADO", "APELLIDO PATERNO", "APELLIDO MATERNO",
        "1ER. NOMBRE", "2DO. NOMBRE", "C. I.", "EXP.", "SEXO",
        "DIRECCIÓN DEL DOMICILIO", "CARGO ACTUAL",
        "UNIDAD DE DESTINO ACTUAL", "FECHA DESTINO ACTUAL",
        "DESTINO ANTERIOR", "NÚMERO DE CELULAR",
        "CORREO ELECTRÓNICO", "OTRA PROFESIÓN (LIC./TÉC.)",
    ]
    COL_WIDTHS = [14, 18, 18, 14, 14, 12, 6, 6, 28, 22, 22, 16, 22, 16, 26, 24]

    h_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    h_fill   = PatternFill('solid', start_color='1F3864')
    h_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_font   = Font(name='Arial', size=9)
    c_align  = Alignment(vertical='center', wrap_text=True)
    side     = Side(style='thin', color='AAAAAA')
    border   = Border(left=side, right=side, top=side, bottom=side)
    alt_fill = PatternFill('solid', start_color='EAF0FB')

    wb = Workbook()
    ws = wb.active
    ws.title = "Personal"

    ws.merge_cells(f'A1:{get_column_letter(len(HEADERS))}1')
    ws['A1'] = 'UTEPPI — REPORTE DE PERSONAL POLICIAL'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='1F3864')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.append([''] * len(HEADERS))
    ws.row_dimensions[2].height = 6

    for col, header in enumerate(HEADERS, 1):
        cell           = ws.cell(row=3, column=col, value=header)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
        cell.border    = border
    ws.row_dimensions[3].height = 36

    for row_idx, p in enumerate(queryset, start=4):
        destino_activo   = p.destinos.filter(activo=True).first()
        destino_anterior = p.destinos.filter(activo=False).order_by('-fecha_inicio').first()

        partes         = (p.nombres or '').split(' ', 1)
        primer_nombre  = partes[0] if partes else ''
        segundo_nombre = partes[1] if len(partes) > 1 else ''

        row_data = [
            p.grado.abreviatura,
            p.apellido_paterno,
            p.apellido_materno,
            primer_nombre,
            segundo_nombre,
            p.ci,
            p.expedido,
            p.get_genero_display(),
            p.direccion_domicilio or '',
            p.cargo_actual or '',
            (destino_activo.unidad_destino.nombre
                if destino_activo and destino_activo.unidad_destino
                else (destino_activo.lugar_destino if destino_activo else '')),
            str(destino_activo.fecha_inicio) if destino_activo else '',
            (f"{destino_anterior.lugar_destino} "
             f"({destino_anterior.fecha_inicio} - "
             f"{destino_anterior.fecha_fin or 'actualidad'})")
                if destino_anterior else '',
            p.telefono_personal or '',
            p.correo_institucional or '',
            p.otra_profesion or '',
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        for col, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=value)
            cell.font      = c_font
            cell.alignment = c_align
            cell.border    = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    for i, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"
    return wb


class ReportePersonalView(OficialAdministrativoRequiredMixin, ListView):
    model               = PersonalPolicial
    template_name       = 'personal/reporte_personal.html'
    context_object_name = 'personal'
    paginate_by         = 50

    def get_queryset(self):
        return self._queryset_filtrado()

    def _queryset_filtrado(self):
        qs = PersonalPolicial.objects.select_related(
            'grado', 'unidad', 'estado_actual'
        ).prefetch_related('destinos__unidad_destino')

        buscar = self.request.GET.get('buscar', '').strip()
        if buscar:
            qs = qs.filter(
                Q(nombres__icontains=buscar)          |
                Q(apellido_paterno__icontains=buscar) |
                Q(apellido_materno__icontains=buscar) |
                Q(ci__icontains=buscar)
            )
        if self.request.GET.get('grado'):
            qs = qs.filter(grado_id=self.request.GET['grado'])
        if self.request.GET.get('unidad'):
            qs = qs.filter(unidad_id=self.request.GET['unidad'])
        if self.request.GET.get('estado'):
            qs = qs.filter(estado_actual_id=self.request.GET['estado'])
        if self.request.GET.get('genero'):
            qs = qs.filter(genero=self.request.GET['genero'])

        return qs.order_by('grado__orden', 'apellido_paterno', 'nombres')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from catalogos.models import Grado, Unidad, TipoEstado
        context.update({
            'grados'         : Grado.objects.filter(activo=True).order_by('orden'),
            'unidades'       : Unidad.objects.filter(activa=True).order_by('nombre'),
            'estados'        : TipoEstado.objects.all(),
            'buscar'         : self.request.GET.get('buscar', ''),
            'grado_sel'      : self.request.GET.get('grado', ''),
            'unidad_sel'     : self.request.GET.get('unidad', ''),
            'estado_sel'     : self.request.GET.get('estado', ''),
            'genero_sel'     : self.request.GET.get('genero', ''),
            'total_filtrado' : self._queryset_filtrado().count(),
            'query_string'   : self.request.GET.urlencode(),
        })
        return context


def exportar_personal_excel(request):
    if not request.user.is_authenticated or not (
        request.user.es_administrador() or request.user.es_oficial_administrativo()
    ):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    qs = PersonalPolicial.objects.select_related(
        'grado', 'unidad', 'estado_actual'
    ).prefetch_related('destinos__unidad_destino')

    buscar = request.GET.get('buscar', '').strip()
    if buscar:
        qs = qs.filter(
            Q(nombres__icontains=buscar)          |
            Q(apellido_paterno__icontains=buscar) |
            Q(apellido_materno__icontains=buscar) |
            Q(ci__icontains=buscar)
        )
    if request.GET.get('grado'):
        qs = qs.filter(grado_id=request.GET['grado'])
    if request.GET.get('unidad'):
        qs = qs.filter(unidad_id=request.GET['unidad'])
    if request.GET.get('estado'):
        qs = qs.filter(estado_actual_id=request.GET['estado'])
    if request.GET.get('genero'):
        qs = qs.filter(genero=request.GET['genero'])

    qs = qs.order_by('grado__orden', 'apellido_paterno', 'nombres')

    # Registrar descarga en bitácora
    registrar_log(
        request, 'OTRO', 'reportes',
        f'Descargó reporte Excel de personal ({qs.count()} registros)',
    )

    wb     = _build_workbook(qs)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import date
    filename = f"reporte_personal_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response