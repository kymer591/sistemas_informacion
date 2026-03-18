import io
from datetime import date

from .utils import registrar_log
from django.db.models import Q
from core.mixins import AdminRequiredMixin
from .models import BitacoraLog

from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.http import HttpResponse
from django.views.generic import ListView

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from personal.models import PersonalPolicial
from catalogos.models import Grado, Unidad, TipoEstado
# Ajusta según el nombre exacto de tu mixin:
from core.mixins import OficialAdministrativoRequiredMixin

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT


HEADERS = [
    "GRADO",
    "APELLIDO PATERNO",
    "APELLIDO MATERNO",
    "1ER. NOMBRE",
    "2DO. NOMBRE",
    "C. I.",
    "EXP.",
    "SEXO",
    "DIRECCIÓN DEL DOMICILIO",
    "CARGO ACTUAL",
    "UNIDAD DE DESTINO ACTUAL",
    "FECHA DESTINO ACTUAL",
    "DESTINO ANTERIOR",
    "NÚMERO DE CELULAR",
    "CORREO ELECTRÓNICO",
    "OTRA PROFESIÓN (LIC./TÉC.)",
]

COL_WIDTHS = [14, 18, 18, 14, 14, 12, 6, 6, 28, 22, 22, 16, 22, 16, 26, 24]


def _aplicar_filtros(qs, params):
    buscar = params.get('buscar', '').strip()
    if buscar:
        qs = qs.filter(
            Q(nombres__icontains=buscar)          |
            Q(apellido_paterno__icontains=buscar) |
            Q(apellido_materno__icontains=buscar) |
            Q(ci__icontains=buscar)
        )
    if params.get('grado'):
        qs = qs.filter(grado_id=params['grado'])
    if params.get('unidad'):
        qs = qs.filter(unidad_id=params['unidad'])
    if params.get('estado'):
        qs = qs.filter(estado_actual_id=params['estado'])
    if params.get('genero'):
        qs = qs.filter(genero=params['genero'])
    return qs


def _build_workbook(queryset):
    h_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    h_fill   = PatternFill('solid', start_color='1F3864')
    h_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_font   = Font(name='Arial', size=9)
    c_align  = Alignment(vertical='center', wrap_text=True)
    side     = Side(style='thin', color='AAAAAA')
    border   = Border(left=side, right=side, top=side, bottom=side)
    alt_fill = PatternFill('solid', start_color='EAF0FB')

    wb = Workbook()
    ws = wb.active
    ws.title = "Personal"

    # Título institucional
    ws.merge_cells(f'A1:{get_column_letter(len(HEADERS))}1')
    ws['A1'] = 'UTEPPI — REPORTE DE PERSONAL POLICIAL'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='1F3864')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila separadora
    ws.append([''] * len(HEADERS))
    ws.row_dimensions[2].height = 6

    # Encabezados fila 3
    for col, header in enumerate(HEADERS, 1):
        cell           = ws.cell(row=3, column=col, value=header)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
        cell.border    = border
    ws.row_dimensions[3].height = 36

    # Datos
    for row_idx, p in enumerate(queryset, start=4):
        destino_activo   = p.destinos.filter(activo=True).first()
        destino_anterior = p.destinos.filter(activo=False).order_by('-fecha_inicio').first()

        partes         = (p.nombres or '').split(' ', 1)
        primer_nombre  = partes[0] if partes else ''
        segundo_nombre = partes[1] if len(partes) > 1 else ''

        row_data = [
            p.grado.abreviatura,
            p.apellido_paterno,
            p.apellido_materno,
            primer_nombre,
            segundo_nombre,
            p.ci,
            p.expedido,
            p.get_genero_display(),
            p.direccion_domicilio or '',
            p.cargo_actual or '',
            (destino_activo.unidad_destino.nombre
                if destino_activo and destino_activo.unidad_destino
                else (destino_activo.lugar_destino if destino_activo else '')),
            str(destino_activo.fecha_inicio) if destino_activo else '',
            (f"{destino_anterior.lugar_destino} "
             f"({destino_anterior.fecha_inicio} - "
             f"{destino_anterior.fecha_fin or 'actualidad'})")
                if destino_anterior else '',
            p.telefono_personal or '',
            p.correo_institucional or '',
            p.otra_profesion or '',
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        for col, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=value)
            cell.font      = c_font
            cell.alignment = c_align
            cell.border    = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    for i, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

    return wb


class ReportePersonalView(OficialAdministrativoRequiredMixin, ListView):
    model               = PersonalPolicial
    template_name       = 'reportes/reporte_personal.html'
    context_object_name = 'personal'
    paginate_by         = 50

    def _base_qs(self):
        return PersonalPolicial.objects.select_related(
            'grado', 'unidad', 'estado_actual'
        ).prefetch_related(
            'destinos__unidad_destino'
        ).order_by('grado__orden', 'apellido_paterno', 'nombres')

    def get_queryset(self):
        return _aplicar_filtros(self._base_qs(), self.request.GET)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'grados'        : Grado.objects.filter(activo=True).order_by('orden'),
            'unidades'      : Unidad.objects.filter(activa=True).order_by('nombre'),
            'estados'       : TipoEstado.objects.all(),
            'buscar'        : self.request.GET.get('buscar', ''),
            'grado_sel'     : self.request.GET.get('grado', ''),
            'unidad_sel'    : self.request.GET.get('unidad', ''),
            'estado_sel'    : self.request.GET.get('estado', ''),
            'genero_sel'    : self.request.GET.get('genero', ''),
            'total_filtrado': self.get_queryset().count(),
            'query_string'  : self.request.GET.urlencode(),
        })
        return context


def exportar_personal_excel(request):
    if not request.user.is_authenticated or not (
        request.user.es_administrador() or request.user.es_oficial_administrativo()
    ):
        raise PermissionDenied

    qs = PersonalPolicial.objects.select_related(
        'grado', 'unidad', 'estado_actual'
    ).prefetch_related(
        'destinos__unidad_destino'
    ).order_by('grado__orden', 'apellido_paterno', 'nombres')

    qs = _aplicar_filtros(qs, request.GET)

    wb     = _build_workbook(qs)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_personal_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# ================================================================
# Agregar al final de reportes/views.py
# ================================================================
# Agregar también este import al inicio de reportes/views.py:
#   from .models import BitacoraLog
# ================================================================



class BitacoraView(AdminRequiredMixin, ListView):
    model               = BitacoraLog
    template_name       = 'reportes/bitacora.html'
    context_object_name = 'logs'
    paginate_by         = 50

    def get_queryset(self):
        qs = BitacoraLog.objects.select_related('usuario')

        buscar = self.request.GET.get('buscar', '').strip()
        if buscar:
            qs = qs.filter(
                Q(usuario__username__icontains=buscar) |
                Q(descripcion__icontains=buscar)       |
                Q(objeto_repr__icontains=buscar)
            )

        if self.request.GET.get('accion'):
            qs = qs.filter(accion=self.request.GET['accion'])

        if self.request.GET.get('modulo'):
            qs = qs.filter(modulo=self.request.GET['modulo'])

        if self.request.GET.get('usuario'):
            qs = qs.filter(usuario_id=self.request.GET['usuario'])

        if self.request.GET.get('fecha_desde'):
            qs = qs.filter(fecha_hora__date__gte=self.request.GET['fecha_desde'])

        if self.request.GET.get('fecha_hasta'):
            qs = qs.filter(fecha_hora__date__lte=self.request.GET['fecha_hasta'])

        return qs

    def get_context_data(self, **kwargs):
        from django.contrib.auth import get_user_model
        Usuario = get_user_model()
        context = super().get_context_data(**kwargs)
        context.update({
            'accion_choices' : BitacoraLog.ACCION_CHOICES,
            'modulo_choices' : BitacoraLog.MODULO_CHOICES,
            'usuarios'       : Usuario.objects.filter(is_active=True).order_by('username'),
            'buscar'         : self.request.GET.get('buscar', ''),
            'accion_sel'     : self.request.GET.get('accion', ''),
            'modulo_sel'     : self.request.GET.get('modulo', ''),
            'usuario_sel'    : self.request.GET.get('usuario', ''),
            'fecha_desde'    : self.request.GET.get('fecha_desde', ''),
            'fecha_hasta'    : self.request.GET.get('fecha_hasta', ''),
            'query_string'   : self.request.GET.urlencode(),
            'total_logs'     : self.get_queryset().count(),
        })
        return context

def exportar_bitacora_pdf(request):
    """
    Exporta la bitácora filtrada a PDF.
    Acceso: Solo Administrador.
    """
    if not request.user.is_authenticated or not request.user.es_administrador():
        raise PermissionDenied
 
    # Aplicar los mismos filtros que la vista de bitácora
    qs = BitacoraLog.objects.select_related('usuario')
 
    buscar = request.GET.get('buscar', '').strip()
    if buscar:
        qs = qs.filter(
            Q(usuario__username__icontains=buscar) |
            Q(descripcion__icontains=buscar)       |
            Q(objeto_repr__icontains=buscar)
        )
    if request.GET.get('accion'):
        qs = qs.filter(accion=request.GET['accion'])
    if request.GET.get('modulo'):
        qs = qs.filter(modulo=request.GET['modulo'])
    if request.GET.get('usuario'):
        qs = qs.filter(usuario_id=request.GET['usuario'])
    if request.GET.get('fecha_desde'):
        qs = qs.filter(fecha_hora__date__gte=request.GET['fecha_desde'])
    if request.GET.get('fecha_hasta'):
        qs = qs.filter(fecha_hora__date__lte=request.GET['fecha_hasta'])
 
    # Registrar la exportación en la propia bitácora
    registrar_log(
        request, 'OTRO', 'reportes',
        f'Exportó bitácora en PDF ({qs.count()} registros)',
    )
 
    # ── Construir PDF ──────────────────────────────────────────────
    response = HttpResponse(content_type='application/pdf')
    filename = f"bitacora_{date.today().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
 
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )
 
    styles = getSampleStyleSheet()
 
    # Estilos personalizados
    estilo_titulo = ParagraphStyle(
        'titulo',
        parent=styles['Title'],
        fontSize=14,
        textColor=colors.HexColor('#1F3864'),
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    estilo_subtitulo = ParagraphStyle(
        'subtitulo',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    estilo_celda = ParagraphStyle(
        'celda',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
    )
 
    elementos = []
 
    # Título
    elementos.append(Paragraph('UTEPPI — BITÁCORA DEL SISTEMA', estilo_titulo))
    elementos.append(Paragraph(
        f'Generado el {date.today().strftime("%d/%m/%Y")} '
        f'por {request.user.username} · {qs.count()} registros',
        estilo_subtitulo,
    ))
 
    # Encabezados de tabla
    encabezados = [
        Paragraph('<b>FECHA Y HORA</b>', estilo_celda),
        Paragraph('<b>USUARIO</b>', estilo_celda),
        Paragraph('<b>ACCIÓN</b>', estilo_celda),
        Paragraph('<b>MÓDULO</b>', estilo_celda),
        Paragraph('<b>DESCRIPCIÓN</b>', estilo_celda),
        Paragraph('<b>OBJETO AFECTADO</b>', estilo_celda),
        Paragraph('<b>IP</b>', estilo_celda),
    ]
 
    # Colores por acción
    COLORES_ACCION = {
        'LOGIN'   : colors.HexColor('#198754'),
        'LOGOUT'  : colors.HexColor('#6c757d'),
        'CREAR'   : colors.HexColor('#0d6efd'),
        'EDITAR'  : colors.HexColor('#ffc107'),
        'ELIMINAR': colors.HexColor('#dc3545'),
        'ERROR'   : colors.HexColor('#212529'),
        'OTRO'    : colors.HexColor('#6c757d'),
    }
 
    # Filas de datos
    filas = [encabezados]
    colores_filas = [None]  # fila de encabezado no tiene color de acción
 
    for log in qs:
        filas.append([
            Paragraph(log.fecha_hora.strftime('%d/%m/%Y %H:%M:%S'), estilo_celda),
            Paragraph(log.usuario.username if log.usuario else '—', estilo_celda),
            Paragraph(log.get_accion_display(), estilo_celda),
            Paragraph(log.modulo.capitalize(), estilo_celda),
            Paragraph(log.descripcion[:120], estilo_celda),
            Paragraph((log.objeto_repr or '—')[:50], estilo_celda),
            Paragraph(log.ip_address or '—', estilo_celda),
        ])
        colores_filas.append(COLORES_ACCION.get(log.accion, colors.grey))
 
    # Anchos de columna (landscape A4 ≈ 25.7cm útil)
    anchos = [3.5*cm, 2.5*cm, 2*cm, 2*cm, 8*cm, 4*cm, 2.5*cm]
 
    tabla = Table(filas, colWidths=anchos, repeatRows=1)
 
    # Estilo base de la tabla
    estilo_tabla = TableStyle([
        # Encabezado
        ('BACKGROUND',   (0, 0), (-1, 0),  colors.HexColor('#1F3864')),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0),  8),
        ('ALIGN',        (0, 0), (-1, 0),  'CENTER'),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUND',(0, 1), (-1, -1), [colors.HexColor('#EAF0FB'), colors.white]),
        # Bordes
        ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#AAAAAA')),
        ('LINEBELOW',    (0, 0), (-1, 0),  1,   colors.HexColor('#1F3864')),
        # Padding
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
        ('LEFTPADDING',  (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])
 
    # Colorear la columna ACCIÓN por tipo
    for i, color_accion in enumerate(colores_filas[1:], start=1):
        if color_accion:
            estilo_tabla.add('TEXTCOLOR', (2, i), (2, i), color_accion)
            estilo_tabla.add('FONTNAME',  (2, i), (2, i), 'Helvetica-Bold')
 
    tabla.setStyle(estilo_tabla)
    elementos.append(tabla)
 
    doc.build(elementos)
    return response
 